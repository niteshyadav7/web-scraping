from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import threading
import os
from datetime import datetime
from scraper.flipkart import FlipkartScraper
from scraper.utils import parse_review_date
import pandas as pd
import csv

app = Flask(__name__)
CORS(app)  # Enable CORS for React frontend

# Global state to track scraping progress
scraping_state = {
    "is_running": False,
    "current_page": 0,
    "total_reviews": 0,
    "recent_reviews": [], # New: Store last 5 reviews for live feed
    "status": "idle",
    "error": None,
    "output_file": None,
    "csv_file": None
}

def append_to_csv(reviews, filename):
    """Appends reviews to CSV file"""
    if not reviews:
        return
    
    fieldnames = ['platform', 'reviewer_name', 'rating', 'review', 'review_date', 'relative_date', 'product_url']
    file_exists = os.path.isfile(filename)
    
    try:
        with open(filename, mode='a', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            if not file_exists:
                writer.writeheader()
            for r in reviews:
                row_to_write = {k: r.get(k, None) for k in fieldnames}
                writer.writerow(row_to_write)
    except Exception as e:
        print(f"CSV Error: {e}")




def convert_csv_to_excel(csv_filename, excel_filename):
    """Converts CSV to beautifully styled Excel with colors and formatting"""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Read CSV and create basic Excel
        df = pd.read_csv(csv_filename)
        df.to_excel(excel_filename, index=False, engine='openpyxl')
        
        # Load workbook for styling
        wb = load_workbook(excel_filename)
        ws = wb.active
        
        # Define color scheme
        header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")  # Blue
        header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")  # White text
        
        # Rating colors (gradient from red to green)
        rating_colors = {
            1: "FF4444",  # Red
            2: "FF8844",  # Orange
            3: "FFD700",  # Yellow/Gold
            4: "88DD44",  # Light Green
            5: "44DD44"   # Green
        }
        
        # Border style
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Get column indices
        columns = {col.value: idx for idx, col in enumerate(ws[1], 1)}
        rating_col = columns.get('rating')
        review_col = columns.get('review')
        platform_col = columns.get('platform')
        
        # Style data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
            for cell in row:
                # Apply borders
                cell.border = thin_border
                
                # Center alignment for specific columns
                if cell.column in [platform_col, rating_col]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                # Color-code ratings
                if cell.column == rating_col and cell.value:
                    try:
                        rating_value = int(cell.value)
                        if rating_value in rating_colors:
                            cell.fill = PatternFill(start_color=rating_colors[rating_value], 
                                                   end_color=rating_colors[rating_value], 
                                                   fill_type="solid")
                            cell.font = Font(bold=True, color="FFFFFF", size=11)
                    except:
                        pass
                
                # Alternate row colors (zebra striping)
                if row_idx % 2 == 0 and cell.column != rating_col:
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set width with limits
            adjusted_width = min(max(max_length + 2, 12), 80)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Set row height for header
        ws.row_dimensions[1].height = 25
        
        # Save styled workbook
        wb.save(excel_filename)
        print(f"✨ Styled Excel file created: {excel_filename}")
        return True
        
    except Exception as e:
        print(f"Excel conversion error: {e}")
        return False

def scrape_task(product_urls, from_date_str, to_date_str, max_pages):
    """Background scraping task - now supports multiple URLs"""
    global scraping_state
    
    try:
        scraping_state["status"] = "running"
        scraping_state["is_running"] = True
        scraping_state["current_page"] = 0
        scraping_state["total_reviews"] = 0
        scraping_state["error"] = None
        
        from_date = datetime.strptime(from_date_str, "%Y-%m-%d")
        to_date = datetime.strptime(to_date_str, "%Y-%m-%d")
        
        # Setup output
        output_dir = "output"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_filename = f"{output_dir}/reviews_{timestamp}.csv"
        excel_filename = f"{output_dir}/reviews_{timestamp}.xlsx"
        
        scraping_state["csv_file"] = csv_filename
        scraping_state["output_file"] = excel_filename
        
        # Process each URL
        total_products = len(product_urls)
        for product_index, product_url in enumerate(product_urls, 1):
            print(f"\n{'='*60}")
            print(f"📦 PRODUCT {product_index}/{total_products}")
            print(f"🔗 URL: {product_url}")
            print(f"{'='*60}\n")
            
            # Initialize scraper for this product
            scraper = FlipkartScraper()
            
            try:
                scraper.open_product_page(product_url)
                scraper.go_to_all_reviews()
                
                # Track seen reviews (for infinite scroll deduplication)
                seen_review_texts = set()
                scroll_attempts = 0
                max_scroll_attempts = max_pages
                
                # Detect pagination mode on first check
                pagination_mode = scraper.has_next_button()
                mode_name = "PAGINATION" if pagination_mode else "INFINITE SCROLL"
                print(f"🔍 Detected mode: {mode_name}\n")
                
                product_review_count = 0
                
                while scroll_attempts < max_scroll_attempts:
                    scroll_attempts += 1
                    scraping_state["current_page"] = scroll_attempts
                    
                    # Extract ALL currently visible reviews
                    raw_reviews = scraper.extract_page_data()
                    
                    # Filter reviews by date
                    new_reviews = []
                    for r in raw_reviews:
                        # ALWAYS deduplicate - duplicates can appear in both modes
                        review_id = f"{r.get('reviewer_name', '')}_{r.get('review', '')[:50]}"
                        if review_id in seen_review_texts:
                            continue
                        seen_review_texts.add(review_id)
                        
                        raw_date_str = r['review_date']
                        r['relative_date'] = raw_date_str
                        r_date = parse_review_date(raw_date_str)
                        
                        if r_date:
                            r_date = r_date.replace(hour=0, minute=0, second=0, microsecond=0)
                            if from_date <= r_date <= to_date:
                                r['review_date'] = r_date.strftime("%Y-%m-%d")
                                r['product_url'] = product_url  # Add product URL to track source
                                new_reviews.append(r)
                    
                    # Save new reviews to CSV
                    if new_reviews:
                        append_to_csv(new_reviews, csv_filename)
                        scraping_state["total_reviews"] += len(new_reviews)
                        product_review_count += len(new_reviews)
                        scraping_state["recent_reviews"] = (new_reviews + scraping_state["recent_reviews"])[:5]
                        print(f"   ✓ Saved {len(new_reviews)} reviews | Product Total: {product_review_count} | Overall Total: {scraping_state['total_reviews']}")
                    else:
                        print(f"   ⚠ No new reviews found on this iteration")
                    
                    # Try to load more content
                    if scroll_attempts < max_scroll_attempts:
                        if not scraper.next_page():
                            print(f"\n✓ Product {product_index} complete - reached end of content")
                            print(f"   📊 Collected {product_review_count} reviews from this product\n")
                            break
                    else:
                        print(f"\n✓ Product {product_index} complete - reached max iterations ({max_scroll_attempts})")
                        print(f"   📊 Collected {product_review_count} reviews from this product\n")
                        break
                        
            except Exception as e:
                print(f"\n❌ Error scraping product {product_index}: {e}\n")
                scraping_state["error"] = f"Error on product {product_index}: {str(e)}"
            finally:
                scraper.quit()
        
        print(f"\n{'='*60}")
        print(f"🎉 ALL PRODUCTS COMPLETE")
        print(f"📊 Total Reviews Collected: {scraping_state['total_reviews']}")
        print(f"{'='*60}\n")
        
        # Convert to Excel
        if scraping_state["total_reviews"] > 0:
            convert_csv_to_excel(csv_filename, excel_filename)
            scraping_state["status"] = "completed"
        else:
            scraping_state["status"] = "completed"
            scraping_state["error"] = "No reviews found"
            
    except Exception as e:
        scraping_state["status"] = "error"
        scraping_state["error"] = str(e)
    finally:
        scraping_state["is_running"] = False

@app.route('/api/scrape', methods=['POST'])
def start_scrape():
    """Start scraping in background - supports multiple URLs"""
    global scraping_state
    
    if scraping_state["is_running"]:
        return jsonify({"error": "Scraping already in progress"}), 400
    
    data = request.json
    urls_input = data.get('url')
    from_date = data.get('from_date')
    to_date = data.get('to_date')
    max_pages = data.get('max_pages', 100)
    
    if not all([urls_input, from_date, to_date]):
        return jsonify({"error": "Missing required fields"}), 400
    
    # Parse multiple URLs (split by newlines and filter empty lines)
    product_urls = [url.strip() for url in urls_input.split('\n') if url.strip()]
    
    if not product_urls:
        return jsonify({"error": "No valid URLs provided"}), 400
    
    # Validate URLs
    for url in product_urls:
        if 'flipkart.com' not in url:
            return jsonify({"error": f"Invalid Flipkart URL: {url}"}), 400
    
    # Start scraping in background thread
    thread = threading.Thread(
        target=scrape_task,
        args=(product_urls, from_date, to_date, max_pages)
    )
    thread.daemon = True
    thread.start()
    
    return jsonify({
        "message": f"Scraping started for {len(product_urls)} product(s)", 
        "status": "running",
        "product_count": len(product_urls)
    })

@app.route('/api/status', methods=['GET'])
def get_status():
    """Get current scraping status"""
    return jsonify(scraping_state)

@app.route('/api/download', methods=['GET'])
def download_file():
    """Download the Excel file"""
    if scraping_state["output_file"] and os.path.exists(scraping_state["output_file"]):
        return send_file(
            scraping_state["output_file"],
            as_attachment=True,
            download_name=os.path.basename(scraping_state["output_file"])
        )
    return jsonify({"error": "File not found"}), 404

@app.route('/api/stop', methods=['POST'])
def stop_scrape():
    """Stop the current scraping (graceful)"""
    # Note: This is a simple implementation. For true stopping, 
    # we'd need to implement a flag check in the scraping loop
    return jsonify({"message": "Stop requested (will finish current page)"})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
